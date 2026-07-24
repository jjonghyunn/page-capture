# page_capture_260724_v3.5.py
# 2026-04-17  Jonghyun Park w/ Claude  — v2.0 초기 버전
# 2026-04-20  Jonghyun Park w/ Claude  — v2.1 is_error_page 다국어 에러 감지 강화 + /common/404/ + Chrome ERR 감지
# 2026-04-29  Jonghyun Park w/ Claude  — v2.2 filename에 OUTPUT_DIR 변수 사용 + raw string 적용 + 파일명 정리(두 번째 날짜=캠페인 날짜 제거)
# 2026-05-22  Jonghyun Park w/ Claude  — v2.3 주석/예시 URL sanitize + 도메인 매칭 로직 상수화 (TARGET_DOMAIN / TARGET_DOMAIN_CN / TARGET_BRAND_KEYWORD) + 설정 상수 파일 상단으로 이동
# 2026-06-18  Jonghyun Park w/ Claude  — v2.4 캡처를 ThreadPoolExecutor 로 병렬화 (MAX_WORKERS) + URL 목록 상단 상수(URLS)로 이동
# 2026-06-19  Jonghyun Park w/ Claude  — v2.4 MAX_WORKERS 사양별 권장값 주석 보강
# 2026-07-06  Jonghyun Park w/ Claude  — v2.5 is_error_page 오탐 수정: aiscPrivateError 는 is_displayed 로, ERR_ 문자열은 title 빈 경우로 한정 (정상 페이지가 error_page 로 skip 되던 문제)
# 2026-07-08  Jonghyun Park w/ Claude  — v2.6 perf log 로 메인 문서 HTTP status 감지 추가 (비영어 404 가 is_error_page 를 빠져나가 캡처되던 문제)
# 2026-07-10  Jonghyun Park w/ Claude  — v2.7 perf log 활성화 후 --headless=new 가 빈 흰 창을 띄우는 문제 → --window-position 으로 창을 화면 밖으로 이동
# 2026-07-10  Jonghyun Park w/ Claude  — v2.8 PC 캡처 분기의 미정의 호출 is_sec_path → is_hq_path 로 수정(매 PC 캡처 NameError 로 skip 되던 문제) + OUTPUT_DIR 기본값을 captures 로 정리
# 2026-07-14  Jonghyun Park w/ Claude  — v2.9 unknown(soft-404) 페이지 skip 추가: 메인 도메인이 존재하지 않는 경로에 HTTP 200 + 홈 fallback 을 주는 경우 <meta property="og:url"> 이 비었거나(EMPTY) 아예 없음(MISSING) → is_unknown_page 로 감지해 skip (리다이렉트/HTTP status/기존 error 마커로 안 잡히던 200 케이스)
# 2026-07-20  Jonghyun Park w/ Claude  — v3.0 [버그] ① 출력 폴더 생성을 캡처 시작 전으로 이동(폴더가 없으면 PNG/MHTML 저장이 전량 실패하던 문제) ② URL 주석 필터를 strip 후 판정(들여쓴 '#' 줄이 URL 로 유입) ③ Chrome 생성을 try 안으로(기동 실패 시 결과 집계에서 통째로 누락) ④ bare except → except Exception (Ctrl+C 중단 가능) ⑤ page load/script timeout 추가(멈춘 서버가 워커를 무한 점유하는 것 방지)
# 2026-07-20  Jonghyun Park w/ Claude  — v3.0 [결과기록] 결과를 (url, device) 단위 result_*.csv 로 기록 — 기존 txt 3종은 URL 단위라 'PC 는 정상인데 MO 만 실패' 를 구분 못 했고 error/timeout 은 어디에도 안 남아 재실행 대상 파악이 불가능했다
# 2026-07-20  Jonghyun Park w/ Claude  — v3.0 [로그인화면 skip] 쿠키·팝업 처리 중 JS 로 늦게 로그인 페이지로 이동하는 경우 리다이렉트 판정(get 직후 1회)을 빠져나가 로그인 화면이 캡처물로 저장됐다 → 저장 직전 URL 재확인 + SKIP_URL_KEYWORDS (RECHECK_URL_BEFORE_SAVE)
# 2026-07-20  Jonghyun Park w/ Claude  — v3.0 [속도] 워커별 Chrome 재사용(REUSE_DRIVER) + 이어하기(SKIP_IF_EXISTS: 같은 날 PNG+MHTML 이 있으면 건너뜀) + 일시적 실패만 재시도(RETRY_COUNT) + MHTML 크기 검증(mhtml_failed)
# 2026-07-20  Jonghyun Park w/ Claude  — v3.0 [리다이렉트 정규화] 후행 슬래시만 무시하던 비교를 #fragment·추적파라미터(utm_*/gclid)·www·대소문자·기본포트·http↔https·퍼센트인코딩·쿼리순서까지 무시하도록 변경 (정상 페이지가 리다이렉트로 오탐돼 skip 되던 문제)
# 2026-07-20  Jonghyun Park w/ Claude  — v3.0 [CDP 전체캡처] Page.captureScreenshot(captureBeyondViewport) 경로를 추가했으나 스크롤로 띄운 지연 로딩 이미지가 렌더되기 전에 찍혀 이미지가 빠진 캡처가 나옴 → 기본 OFF(USE_CDP_FULLPAGE=False), 이미지 로딩 완료 대기를 넣기 전까지 켜지 말 것
# 2026-07-21  Jonghyun Park w/ Claude  — v3.1 SKIP_URL_KEYWORDS 에 /registration 추가 — 메일 인증 게이트 페이지가 URL 에 /login 이 없어 로그인 skip 을 통과해 캡처물로 저장되던 문제
# 2026-07-22  Jonghyun Park w/ Claude  — v3.2 [MO 스티칭 재작성] 겹침 보정이 CSS px 값을 device px 자리에 써서 이음새마다 내용이 중복되고 캔버스가 커진 만큼 하단이 검정으로 남던 버그를 고쳤다(실측 75,492px 페이지에서 중복·검정 약 12%). 이제 각 컷을 "실제 스크롤 위치 × 배율" 자리에 그대로 붙인다 + NEUTRALIZE_STICKY(스티칭 동안 position:fixed/sticky 를 static 으로, 저장 전 원복) + 결과 픽셀에서 스티키 바 높이를 직접 재서 잘라내는 _sticky_band + MOBILE_STITCH_OVERLAP / MOBILE_MAX_SHOTS 상수화
# 2026-07-22  Jonghyun Park w/ Claude  — v3.2 [타임아웃] ① PAGE_LOAD_TIMEOUT 60 → 120 (무거운 페이지는 단독 실행에서도 90초를 넘겨 60초에 걸려 통째로 누락됐다) ② CDP_TIMEOUT 신설 — set_page_load_timeout/set_script_timeout 은 CDP 에 적용되지 않아 Page.captureSnapshot 이 무응답이면 워커가 영구 대기하고 그 future 하나 때문에 결과 CSV 기록까지 막혔다(실제 4시간 점유 사고) → cdp_with_timeout 으로 감싸고 초과 시 그 건만 mhtml_failed 로 포기
# 2026-07-22  Jonghyun Park w/ Claude  — v3.2 [결과 CSV] 캡처 1건이 끝날 때마다 result_*.csv 에 append. 예전엔 전부 끝난 뒤 한 번에 써서, 중간에 멈추면 그날 판정 결과가 통째로 날아갔다. 완주하면 마지막에 입력 URL 순서로 정렬해 다시 쓴다
# 2026-07-22  Jonghyun Park w/ Claude  — v3.3 [hang 원인 제거] driver.quit() 이 응답 없는 chromedriver 를 상대로 타임아웃 없는 HTTP 요청(send_remote_shutdown_command → is_connectable)을 걸어 영구 대기하던 문제 — 스택 덤프로 확인된 실제 hang 지점. 별도 스레드로 quit 하고 QUIT_TIMEOUT 초과 시 chromedriver 프로세스를 직접 종료
# 2026-07-22  Jonghyun Park w/ Claude  — v3.3 [진단] 파일 로그(run_*.log)로 출력 미러링 + STACK_DUMP_INTERVAL 마다 전체 스레드 스택 덤프. 스케줄러가 pythonw 로 돌리면 콘솔 출력이 버려져 hang 원인 추적이 불가능했다 (이 로그가 위 원인을 하루 만에 특정했다)
# 2026-07-22  Jonghyun Park w/ Claude  — v3.3 [신뢰성] ① 태스크 하드 데드라인(TASK_DEADLINE_MO/PC) — 감시 스레드가 상한 초과 워커의 드라이버를 끊어 전체 실행이 물리는 것을 막는다 ② 시작 시 이전 실행 잔재(headless chrome/chromedriver) 정리 — 강제 종료 시 Chrome 이 남아 메모리를 물고 있었다
# 2026-07-22  Jonghyun Park w/ Claude  — v3.3 [속도] HTTP 사전 필터(PREFILTER_HTTP) — 404/5xx 는 브라우저 없이 확정. 실측상 워커시간의 58%가 죽은 URL 확인에 쓰였고 대상 URL 의 84%가 PC·MO 양쪽 死였다. 403 은 봇 차단일 수 있어 제외, soft-404 는 기존 og:url 검사가 계속 담당
# 2026-07-22  Jonghyun Park w/ Claude  — v3.3 [산출물] 일일 리포트 daily_report.xlsx — 최신 날짜가 항상 B~D열(PC/MO/이슈)이고 과거 날짜는 오른쪽으로 밀린다. 같은 날 재실행은 덮어쓰기 + MO 하단 흰 여백 트림(TRIM_TRAILING_BLANK)
# 2026-07-23  Jonghyun Park w/ Claude  — v3.4 [크래시] STACK_DUMP_INTERVAL 기본 OFF. 이 기능을 켠 뒤 access violation(0xC0000005) 이 2회 발생했고 둘 다 "덤프 출력 도중"이었다. 하드 크래시는 finally 도 안 돌아, 전체 태스크의 절반쯤에서 죽으면 큐 뒤쪽이 통째로 결번되고 산출물(리포트·skip txt)도 하나도 안 남는다 — 실제로 하루치 캡처의 27%가 그렇게 빠졌다
# 2026-07-23  Jonghyun Park w/ Claude  — v3.4 [산출물] 날짜별 result CSV·run 로그 누적을 폐지하고 고정명·언더바 접두로 통일 — _daily_report.xlsx / _result_latest.csv / _run_latest.log / _skipped_*.txt (RESULT_CSV_NAME·RUN_LOG_NAME·SKIPPED_TXT_NAMES). 이력은 일일 리포트 한 파일이 갖고, 폴더 맨 위에 모여 보인다
# 2026-07-23  Jonghyun Park w/ Claude  — v3.4 [리포트 내구성] write_daily_report 가 메모리 rows 대신 result CSV 를 읽도록 바꾸고 REPORT_FLUSH_EVERY(기본 50) 건마다 중간 저장. 예전엔 완주해야만 리포트가 생겨, 중간에 죽으면 "어디까지 됐는지" 조차 남지 않았다
# 2026-07-23  Jonghyun Park w/ Claude  — v3.4 [이어하기] already_captured() 신설 — SKIP_IF_EXISTS 가 OUTPUT_DIR 루트만 보던 탓에, 캡처물을 sitecode 폴더로 옮기는 foldering 을 돌린 뒤에는 이어하기가 무력화돼 재실행이 전량 재캡처가 됐다. 이제 {OUTPUT_DIR}/{SITECODE}/ 도 함께 확인한다
# 2026-07-24  Jonghyun Park w/ Claude  — v3.5 [산출물 축소] 폴더에 _daily_report.xlsx 하나만 남긴다: LOG_TO_FILE 기본 OFF(_run_latest.log 미생성) / WRITE_SKIPPED_TXT=False(_skipped_*.txt 미생성, 개수는 콘솔로) / KEEP_RESULT_CSV=False(_result_latest.csv 는 실행 중엔 리포트 입력으로 유지하다 완주 후 삭제). 크래시 시엔 CSV 가 남아 사후 추적 가능
#
# ── 캡처한 URL 확인법 (저장된 .mhtml) ──────────────────────────────
# 저장된 .mhtml 을 텍스트 에디터(메모장 등)로 열면 맨 위 MIME 헤더 2번째 줄
#   Snapshot-Content-Location: <실제 캡처된 URL>
# 이 Chrome 이 캡처 시점에 기록한 메인 문서 URL (쿼리 파라미터까지 보존).
# Subject = 페이지 타이틀, Date = 캡처 시각도 같이 기록됨.
# ⚠ 본문 안의 href(예: AudioEye skip-link href="...#content")는 JS 가 동적 생성한
#   링크라 보통 일치하긴 하나 정식 기록이 아님 — 반드시 위 헤더를 볼 것.
# (이 스크립트는 driver.current_url 이 요청 URL 과 다르면(리다이렉트) 저장 자체를
#  skip 하므로, Snapshot-Content-Location = 요청한 URL 이 보장됨)
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from datetime import datetime
from urllib.parse import urlparse, parse_qsl, urlencode, unquote
import time
from PIL import Image
import io
import re
import os
import json
import csv
import sys
import base64
import threading
import faulthandler
import subprocess
import atexit
from collections import Counter
from concurrent.futures import ThreadPoolExecutor, as_completed
import numpy as np
import requests

# ════ 사용자가 바꿔야 하는 부분 (상단 설정) ════
# 저장 경로
OUTPUT_DIR = r"C:\Users\user_name\Downloads\captures"

# 동시에 띄울 헤드리스 Chrome 개수. 브라우저 1개당 ~300-500MB RAM.
# 한계: ① RAM (워커수 × ~0.5GB) ② CPU 논리프로세서 수(초과분은 효과 없이 task당 시간만 늘어남)
#       ③ 대상 서버 throttle(너무 많이 동시 요청 시 리다이렉트/타임아웃 헛 skip ↑)
# → 무한정 못 올림. 보통 "논리프로세서 수" 와 "여유 RAM ÷ 0.5GB" 중 작은 값 근처가 상한.
#
# 사양별 권장값:
#   • 이 PC (6코어/12논리, RAM 31GB)            → 8~10 권장 (현재 작업 기준 10이 적정 상한)
#   • 16GB RAM / 4코어8논리 노트북              → 4~6
#   • 8GB RAM 저사양                            → 2~4 (그 이상은 스왑으로 오히려 느려짐)
#   • 32GB+ / 8코어16논리 이상 데스크탑          → 12~14
# (다른 앱이 떠 있으면 여유 RAM 이 줄어드니 1~2 낮춰 잡을 것)
#
# ⚠ PREFILTER_HTTP 를 켜면 이 값을 더 보수적으로 잡을 것.
#   사전 필터가 404/5xx 를 걷어내면 남는 태스크가 "전부 실제로 렌더되는 무거운 페이지"가 된다.
#   예전엔 가벼운 404 가 섞여 메모리가 중간중간 숨을 쉬었는데 그 완충이 사라지는 셈이다.
#   MO 스티칭은 컷 이미지를 전부 메모리에 들고 있어(7만 px 페이지면 컷만 1GB 이상)
#   워커가 동시에 무거운 페이지를 잡으면 십수 GB 가 순간적으로 뛴다.
#   (실측: 31GB PC 에서 10 워커로 전체 실행 중 여유 RAM 이 3GB 아래로 떨어져 프로세스가 죽음)
MAX_WORKERS = 4

# 페이지 로딩 제한 시간(초). 초과하면 그 (url, device) 는 timeout 실패로 기록하고 다음 작업으로 넘어간다.
# (미설정 시 서버가 hang 하면 워커 1개가 무한정 물려 전체 진행이 막힘)
# ⚠ 목적은 "느린 페이지 거르기"가 아니라 "멈춘 서버가 워커를 무한 점유하는 것 막기" — 넉넉히 잡을 것.
#   정상 페이지 단독 로드는 2~5초지만, MAX_WORKERS 개를 동시에 띄우면 CPU·대역폭 경합으로
#   같은 페이지가 10초를 훌쩍 넘긴다(10초로 잡았다가 정상 site 다수가 죽는 것을 확인, 2026-07-20).
#   2026-07-21: 60 → 120. US education 처럼 무거운 페이지(캡처 PNG 9MB)가 단독 실행에서도
#   93초가 걸려 60초에 걸려 통째로 누락됐다(그날 어제 대비 1건 누락의 원인).
PAGE_LOAD_TIMEOUT = 120
# execute_script / execute_async_script 제한 시간(초)
SCRIPT_TIMEOUT = 30
# CDP(execute_cdp_cmd) 제한 시간(초).
# ⚠ 위 두 타임아웃은 CDP 명령에 적용되지 않는다 — Page.captureSnapshot 이 무응답이면
#   워커가 영구 대기하고, 그 future 때문에 as_completed 루프가 안 끝나 결과 CSV 도 못 쓴다.
#   (2026-07-21: EE_PC 한 건이 4시간 넘게 물려 전체 실행이 멈춘 사고)
CDP_TIMEOUT = 90

# ── 로그인/인증 화면 캡처 제외 ────────────────────────────────
# 리다이렉트 판정은 driver.get() 직후 1회뿐이라, 쿠키·팝업 처리 중 JS 로 "늦게" 튀는
# 로그인 페이지(/auth/multistore 등)는 판정을 통과한 뒤 이동해 그대로 저장돼 버린다.
# (2026-07-20 확인: 그날 저장분 90개 중 7개가 로그인 화면)
# True 면 저장 직전에 current_url 을 한 번 더 확인해 걸러낸다.
RECHECK_URL_BEFORE_SAVE = True
# 최종 URL 에 아래 조각이 들어가면 캡처하지 않고 skip (result = login_page)
# ⚠ /registration = 학생포털 이메일 인증 게이트(NL/EG). 이름만 registration 일 뿐
#   "학생 메일 주소를 넣으면 접근 링크를 보내준다"는 로그인 화면과 동일해 캡처 가치가 없다.
#   (2026-07-21 확인: NL 은 URL 에 /login 이 없어 위 4개 키워드를 통과해 그대로 저장됐다)
SKIP_URL_KEYWORDS = ["/auth/", "/login", "/signin", "/sign-in", "/registration"]

# ── 리다이렉트 판정 정규화 ────────────────────────────────────
# 요청 URL 과 최종 URL 을 비교할 때 "사실상 같은 페이지"인 차이는 무시한다.
# (예전엔 후행 슬래시만 무시해서, #fragment 나 utm_* 가 붙기만 해도 리다이렉트로 보고 skip 했다)
# 아래 파라미터는 페이지 내용과 무관한 추적용이라 비교에서 제외한다.
TRACKING_PARAMS = {
    "utm_source", "utm_medium", "utm_campaign", "utm_term", "utm_content", "utm_id",
    "gclid", "fbclid", "msclkid", "igshid", "mc_cid", "mc_eid", "_ga", "_gl",
    "cid", "CID", "campaign", "src", "source",
}
# www. 유무 차이를 무시할지 (www.example.com ↔ example.com 은 같은 페이지로 본다)
IGNORE_WWW_IN_COMPARE = True

# ── 속도 옵션 ────────────────────────────────────────────────
# 워커(스레드)마다 Chrome 을 1개만 띄우고 URL 이 바뀌어도 재사용한다.
# 끄면 예전처럼 (url, device) 마다 새로 기동 — 기동에만 건당 3~10초가 든다.
REUSE_DRIVER = True
# 이미 같은 이름의 PNG+MHTML 이 있으면 건너뛴다 (중단 후 이어하기).
# 파일명에 날짜(%m%d)가 들어가므로 "같은 날 재실행분만" 건너뛴다.
SKIP_IF_EXISTS = True
# 캡처 실패(timeout/error) 시 재시도 횟수. 404·리다이렉트 등 판정 결과는 재시도하지 않는다.
RETRY_COUNT = 1
# 전체 페이지 캡처를 CDP(captureBeyondViewport)로 한다.
# 끄면 기존 방식(PC=창 늘리기 / MO=스크롤 스티칭) 사용.
# ⚠ 현재 False 가 정답 — CDP 는 스크롤로 띄운 레이지 로딩 이미지가 렌더되기 전에 찍혀
#   제품 이미지가 통째로 빠진 캡처가 나온다(2026-07-20 MO 캡처에서 확인).
#   높이는 CDP 쪽이 정확하지만(중복 접합·하단 여백 없음) 내용이 비므로 쓸 수 없다.
#   → 이미지 로딩 완료를 기다리는 로직을 넣기 전까지는 켜지 말 것.
USE_CDP_FULLPAGE = False
# CDP 캡처 최대 높이(px). 너무 긴 페이지는 Chrome 렌더 한계·메모리 문제로 잘라낸다.
MAX_CAPTURE_HEIGHT = 30000
# MHTML 최소 크기(byte). 이보다 작으면 저장 실패로 보고 result=mhtml_failed 로 남긴다.
MIN_MHTML_BYTES = 20000

# ── 실행 로그 / 잔재 정리 ────────────────────────────────────
# 작업 스케줄러는 pythonw 로 돌아 콘솔 출력이 통째로 버려진다. 그래서 hang 이 나도
# "어디서 멈췄는지" 를 사후에 알 수 없었다(2026-07-21·22). 파일로 같이 남긴다.
# ⚠ 2026-07-24: 산출물을 _daily_report.xlsx 하나로만 남기려고 기본 OFF 로 전환.
#   끄면 pythonw 스케줄 실행에서 hang/크래시 원인 추적용 콘솔 로그가 사라진다 — hang 을
#   다시 파야 하면 True 로 되돌릴 것. (데드라인 감시 TASK_DEADLINE_* 는 그대로 동작한다.)
LOG_TO_FILE = False
# 실행 로그 파일명 — 날짜별로 쌓지 않고 매 실행 덮어쓴다(언더바 접두 = 폴더 맨 위 정렬).
RUN_LOG_NAME = "_run_latest.log"
# 이 간격(초)마다 전체 스레드 스택을 로그에 덤프. 0 이면 끔.
# ⚠ 2026-07-23: 기본 OFF. 이 기능을 켠 뒤 관측된 access violation(0xC0000005) 2건이
#   모두 "덤프 출력 도중" 발생했다(전체 태스크의 절반쯤에서 프로세스가 통째로 사망 → 그날
#   캡처가 큐 뒤쪽부터 통째로 결번). 프로세스가 죽으면 finally 도 안 돌아 뒷정리·산출물이
#   전부 날아간다. hang 추적은 run 로그 + 데드라인 감시로 충분하다.
#   (hang 을 다시 파야 할 때만 일시적으로 켤 것)
STACK_DUMP_INTERVAL = 0
# 시작할 때 이전 실행이 남긴 headless Chrome / chromedriver 를 정리할지.
# ⚠ 임시 프로필(scoped_dir)·--headless 인 것만 죽인다 — 사용자가 쓰는 Chrome 은 건드리지 않는다.
#   (PT3H 강제 종료 시 파이썬만 죽고 Chrome 100여 개가 남아 메모리를 물고 있었다)
SWEEP_STALE_CHROME = True

# ── 태스크 하드 데드라인 ─────────────────────────────────────
# (url, device) 하나가 이 시간을 넘기면 그 워커의 chromedriver 를 강제로 끊어
# 태스크를 timeout 으로 확정한다. 개별 명령 타임아웃(PAGE_LOAD/SCRIPT/CDP)은
# "명령 하나"만 감시하므로, 명령 사이에서 늘어지는 경우를 못 잡는다.
# ⚠ 실측(2026-07-22) 최대 401초(MO)/278초(PC) → 넉넉히 잡는다. 0 이면 끔.
TASK_DEADLINE_MO = 900        # 초
TASK_DEADLINE_PC = 600        # 초
DEADLINE_CHECK_INTERVAL = 15  # 감시 주기(초)
# driver.quit() 응답 대기 상한(초). 넘으면 chromedriver 프로세스를 직접 죽인다.
# ⚠ selenium 의 quit 경로(send_remote_shutdown_command → is_connectable)는 타임아웃이 없어
#   chromedriver 가 응답을 멈추면 영구 대기한다 — 2026-07-22 스택 덤프로 확인된 실제 hang 지점.
QUIT_TIMEOUT = 30

# ── HTTP 사전 필터 ───────────────────────────────────────────
# 브라우저를 띄우기 전에 URL 상태코드를 먼저 확인해, 확실히 죽은 것은 Chrome 없이 확정한다.
# (2026-07-22 실측: 워커시간의 58%가 404 확인에 쓰였고, 975개 중 822개가 PC·MO 양쪽 死)
PREFILTER_HTTP = True
# 이 상태코드면 캡처하지 않고 error_page 로 확정한다.
# ⚠ 403 은 넣지 말 것 — 봇 차단일 수 있어 브라우저로는 정상 렌더되는 경우가 있다.
PREFILTER_DEAD_STATUS = {404, 410, 500, 502, 503, 504}
PREFILTER_WORKERS = 20        # 사전 확인 동시 요청 수
PREFILTER_TIMEOUT = 15        # 초
# 사전 확인용 User-Agent (모바일/데스크탑 공통 — 상태코드만 보므로 데스크탑으로 통일)
PREFILTER_UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                "(KHTML, like Gecko) Chrome/120.0 Safari/537.36")

# ── 일일 리포트 ──────────────────────────────────────────────
# 날짜별 결과를 한 파일에 누적한다. 최신 날짜가 항상 왼쪽(B~D열)에 오고
# 과거 날짜는 오른쪽으로 밀리므로, 열어서 A~D열만 보면 오늘 상태를 알 수 있다.
# ⚠ 날짜별 result CSV 를 쌓는 대신 이 리포트 하나로 이력을 본다(2026-07-23 전환).
DAILY_REPORT = True
DAILY_REPORT_NAME = "_daily_report.xlsx"
# 하루당 열 개수(PC 결과 / MO 결과 / 이슈) — 구조를 바꾸려면 write_daily_report 도 같이 볼 것
DAILY_REPORT_COLS = 3
# 결과 CSV 파일명 — 실행 중 건별로 append 되는 안전망. 날짜별로 쌓지 않고 매 실행 덮어쓴다.
# (이력은 일일 리포트가 갖는다)
# ⚠ 이 CSV 는 write_daily_report 의 입력 원본이라 실행 중에는 반드시 필요하다(메모리 rows 가
#   아니라 디스크 CSV 를 읽는다 — access violation 대비, 2026-07-23). 그래서 "안 만들기"가 아니라
#   완주 후 삭제하는 방식으로 폴더에서 치운다(KEEP_RESULT_CSV=False). 크래시 시엔 남아 사후 추적용.
RESULT_CSV_NAME = "_result_latest.csv"
# 완주(정상 종료) 후 결과 CSV 를 남길지. False = 리포트 갱신까지 마친 뒤 삭제(폴더엔 리포트만 남음).
# 크래시로 중단되면 이 값과 무관하게 CSV 가 남아 그 시점까지의 판정을 볼 수 있다.
KEEP_RESULT_CSV = False
# skip 사유별 txt 파일을 따로 남길지. False = 안 남김(같은 정보가 리포트 '이슈' 열에 이미 있다).
WRITE_SKIPPED_TXT = False
# skip 사유별 txt 파일명 — 역시 고정명(누적 X). 같은 정보가 리포트 '이슈' 열에도 남는다.
SKIPPED_TXT_NAMES = {
    "redirect":     "_skipped_redirect.txt",
    "error_page":   "_skipped_error_page.txt",
    "login_page":   "_skipped_login_page.txt",
    "unknown_page": "_skipped_unknown_page.txt",
}
# 캡처 N건마다 리포트를 중간 저장한다(0 이면 완주 시에만).
# ⚠ access violation 같은 하드 크래시는 finally·atexit 도 실행되지 않는다 — 프로세스가 죽어도
#   그 시점까지의 결과가 리포트에 남으려면 실행 중에 디스크로 내려두는 수밖에 없다(2026-07-23).
REPORT_FLUSH_EVERY = 50

# ── 모바일 전체페이지 스티칭 ──────────────────────────────────
# 이음새에서 겹쳐 찍을 양(CSS px). 스크롤은 (뷰포트 높이 - 이 값) 만큼 내려간다.
MOBILE_STITCH_OVERLAP = 100
# 스티칭 전에 position:fixed / sticky 요소를 static 으로 바꿀지.
# 안 바꾸면 sticky 헤더·하단바가 스크롤 위치마다 다시 찍혀 이음새마다 반복 노출된다.
# (2026-07-21 실측: 높이 75,492px 중 중복·검정이 약 12%)
NEUTRALIZE_STICKY = True
# 스티칭 장수 상한 — 무한 스크롤 페이지에서 메모리 폭주를 막는다.
# ⚠ 스티키 바가 크면 겹침이 늘어 컷 수도 함께 늘어난다(겹침 210px 기준 25,000 CSS px 페이지 ≈ 40장).
#   상한에 걸리면 페이지 아래쪽이 조용히 잘리므로 넉넉히 두고, 걸릴 때는 경고를 남긴다.
MOBILE_MAX_SHOTS = 120
# 이어붙인 뒤 하단에 남는 흰 여백을 잘라낼지 (마지막 컷 뷰포트가 페이지 끝을 넘어서 생긴다)
TRIM_TRAILING_BLANK = True
# 트림할 때 콘텐츠 아래로 남겨둘 여백(px) — 0 으로 바짝 자르면 답답해 보인다
TRIM_KEEP_MARGIN = 40
# get() 후 렌더 안정화 대기 상한(초). readyState 가 complete 면 더 안 기다린다.
PAGE_SETTLE_TIMEOUT = 8

# ── 대상 도메인 ──────────────────────────────────────────────
# ⚠ 도메인 문자열을 함수 본문에 직접 쓰지 말 것.
#   여기 상수만 바꾸면 다른 브랜드·사이트에도 그대로 쓸 수 있고,
#   공개 저장소에 올릴 때 이 몇 줄만 교체하면 로직이 안 깨진다.
#   (함수 안에 박아두면 문자열 치환 시 조건이 조용히 항상 False 가 되어
#    get_site_type / is_hq_path / is_unknown_page 가 티 없이 오작동한다)
TARGET_DOMAIN = "example.com"                      # 메인 도메인 (shop./www. 등 서브도메인 포함해 endswith 로 판정)
TARGET_DOMAIN_CN = ("example.com.cn", "example.cn")  # 중국 등 별도 도메인 → sitecode 를 CN 으로
TARGET_BRAND_KEYWORD = "example"                           # 위에 안 걸리는 브랜드 host 판정용 (host 안에 포함되는지)
# soft-404(og:url) 검사 대상 host — 메인 도메인의 루트/www 만 (shop.* 등은 원래 og:url 이 없어 오탐)
OG_CHECK_HOSTS = {TARGET_DOMAIN, f"www.{TARGET_DOMAIN}"}
# 국가 코드 대신 이 경로로 시작하면 본사(HQ) 페이지 — 렌더가 느려 추가 대기가 필요
HQ_SITE_CODE = "hq"

# ── 뷰포트 / User-Agent ──────────────────────────────────────
DESKTOP_VIEWPORT = (1920, 1080)
MOBILE_VIEWPORT = (390, 844)
MOBILE_SCALE = 3          # 모바일 deviceScaleFactor (레티나 배율)
MOBILE_UA = ('Mozilla/5.0 (iPhone; CPU iPhone OS 15_0 like Mac OS X) AppleWebKit/605.1.15 '
             '(KHTML, like Gecko) Version/15.0 Mobile/15E148 Safari/604.1')

# 캡처할 URL 목록 (한 줄에 하나, # 로 시작하면 주석 처리)
# https:// 필수 작성 필요
URLS = """
https://www.example.com/nz/offer/campaign-name-gift-ideas
https://www.example.com/vn/offer/campaign-name
"""
# ═══════════════════════════════════════════

# =========================
# 실행 로그 / 잔재 정리
# =========================
class _Tee:
    """stdout 을 화면과 파일에 동시에 쓴다. pythonw 로 돌 때는 파일에만 남는다."""
    def __init__(self, stream, fh):
        self.stream, self.fh = stream, fh
        self._lock = threading.Lock()

    def write(self, s):
        with self._lock:
            if self.stream is not None:
                try:
                    self.stream.write(s)
                except Exception:
                    pass
            try:
                self.fh.write(s)
                self.fh.flush()      # hang 나도 직전까지의 로그가 남아야 한다
            except Exception:
                pass

    def flush(self):
        for t in (self.stream, self.fh):
            try:
                if t is not None:
                    t.flush()
            except Exception:
                pass


def start_file_log(ts):
    """OUTPUT_DIR/{RUN_LOG_NAME} 로 출력을 미러링하고, 필요하면 주기적 스택 덤프를 건다.

    로그는 실행마다 새로 쓴다("w") — 날짜별로 쌓지 않는 대신 이력은 일일 리포트가 갖는다.
    """
    if not LOG_TO_FILE:
        return None
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    path = f"{OUTPUT_DIR}/{RUN_LOG_NAME}"
    fh = open(path, "w", encoding="utf-8")
    sys.stdout = _Tee(getattr(sys, "__stdout__", None), fh)
    sys.stderr = _Tee(getattr(sys, "__stderr__", None), fh)
    atexit.register(lambda: fh.flush())
    print(f"📝 실행 로그: {path}")
    if STACK_DUMP_INTERVAL:
        # 멈춘 지점을 알려주는 유일한 단서 — 파일 핸들로 직접 덤프한다
        faulthandler.enable(file=fh)
        faulthandler.dump_traceback_later(STACK_DUMP_INTERVAL, repeat=True, exit=False, file=fh)
        print(f"🩺 {STACK_DUMP_INTERVAL}초마다 스레드 스택 덤프 (hang 원인 추적용)")
    return path


# 이전 실행 잔재만 고르는 조건 — 임시 프로필이거나 headless 인 Chrome + chromedriver 전부.
# 사용자가 직접 쓰는 Chrome 은 자기 프로필(User Data)로 뜨므로 여기에 걸리지 않는다.
_SWEEP_PS = (
    "$c = Get-CimInstance Win32_Process -Filter \"Name='chrome.exe'\" | "
    "Where-Object { $_.CommandLine -match '--headless' -or $_.CommandLine -match 'Temp\\\\scoped_dir' }; "
    "$n = ($c | Measure-Object).Count; "
    "$c | ForEach-Object { Stop-Process -Id $_.ProcessId -Force -ErrorAction SilentlyContinue }; "
    "$d = Get-Process chromedriver -ErrorAction SilentlyContinue; "
    "$m = ($d | Measure-Object).Count; "
    "$d | Stop-Process -Force -ErrorAction SilentlyContinue; "
    "Write-Output \"$n $m\""
)


def sweep_stale_chrome():
    """이전 실행이 강제 종료되며 남긴 headless Chrome / chromedriver 를 정리한다.

    ⚠ 반드시 캡처 시작 "전"에만 부른다 — 실행 중에 부르면 자기 드라이버를 죽인다.
    """
    if not SWEEP_STALE_CHROME or os.name != "nt":
        return
    try:
        out = subprocess.run(["powershell", "-NoProfile", "-Command", _SWEEP_PS],
                             capture_output=True, text=True, timeout=60).stdout.strip()
        chrome_n, driver_n = (int(x) for x in out.split()[:2])
        if chrome_n or driver_n:
            print(f"🧹 이전 실행 잔재 정리: chrome {chrome_n}개 / chromedriver {driver_n}개")
    except Exception as e:
        print(f"  ⚠️ 잔재 정리 건너뜀: {type(e).__name__}: {e}")


# =========================
# 파일명 / 경로 관련 유틸
# =========================
def safe_filename(name: str) -> str:
    name = re.sub(r'[\\/*?:"<>|]+', '_', name)
    name = re.sub(r'\s+', ' ', name).strip()
    return name[:120] if len(name) > 120 else name

def extract_last_slug(parts):
    for seg in reversed(parts):
        s = seg.strip()
        if not s:
            continue
        s = s.split('?')[0].split('#')[0]
        s = re.sub(r'\.html?$','', s, flags=re.I)
        return s
    return 'page'

def normalize_url_for_compare(u: str) -> str:
    """리다이렉트 판정용 정규화. '사실상 같은 페이지'면 같은 문자열이 되도록 만든다.

    무시하는 차이: 대소문자(scheme/host), 기본 포트, www., #fragment,
                   후행 슬래시, 추적 파라미터(utm_* 등), 쿼리 순서, 퍼센트 인코딩.
    경로 자체가 달라지는 진짜 리다이렉트(/auth/multistore 등)는 그대로 다르게 남는다.
    """
    if not u:
        return ""
    try:
        p = urlparse(u.strip())
        host = (p.hostname or "").lower()
        if IGNORE_WWW_IN_COMPARE and host.startswith("www."):
            host = host[4:]
        # 기본 포트는 표기 차이일 뿐이라 제거
        port = p.port
        if port and not ((p.scheme == "https" and port == 443) or (p.scheme == "http" and port == 80)):
            host = f"{host}:{port}"
        path = unquote(p.path).rstrip("/")
        # 추적 파라미터 제거 + 나머지는 순서 무관하게 정렬
        qs = [(k, v) for k, v in parse_qsl(p.query, keep_blank_values=True)
              if k.lower() not in {t.lower() for t in TRACKING_PARAMS}]
        query = urlencode(sorted(qs))
        # scheme 은 http/https 차이를 무시 (같은 페이지의 프로토콜 승격일 뿐)
        return f"{host}{path}?{query}" if query else f"{host}{path}"
    except Exception:
        return (u or "").rstrip("/")

def get_site_type(url: str) -> str:
    parsed = urlparse(url)
    host = parsed.netloc.lower().replace('www.', '')
    parts = [p for p in parsed.path.split('/') if p]

    if host in TARGET_DOMAIN_CN:
        if len(parts) <= 1:
            return 'home'
        return extract_last_slug(parts)

    if host.endswith(TARGET_DOMAIN):
        if len(parts) <= 1:
            return 'home'
        return extract_last_slug(parts)

    if TARGET_BRAND_KEYWORD in host:
        if len(parts) == 0:
            return 'home'
        return extract_last_slug(parts)

    st = extract_last_slug(parts) or 'home'
    return st if st else 'home'

def get_page_info(url):
    """사이트코드, 페이지명, sitetype 추출"""
    parsed = urlparse(url)
    domain = parsed.netloc.replace('www.', '')
    path = parsed.path.strip('/').split('/')

    if any(d in domain for d in TARGET_DOMAIN_CN):
        sitecode = 'CN'
        page_path = '_'.join(p for p in path if p) if path else 'page'

    else:
        sitecode = path[0].upper() if len(path) > 0 and path[0] else 'GLOBAL'
        page_path = '_'.join(p for p in path[1:] if p) if len(path) > 1 else 'page'

    sitetype = get_site_type(url)
    return sitecode, page_path, sitetype

# =========================
# 대기/팝업 관련 유틸
# =========================
def close_popups(driver):
    try:
        close_buttons = [
            "button[aria-label*='close']", "button.close", ".close-button",
            "[class*='close']", "button[title*='Close']", "svg[class*='close']",
        ]
        for selector in close_buttons:
            try:
                buttons = driver.find_elements(By.CSS_SELECTOR, selector)
                for btn in buttons:
                    if btn.is_displayed():
                        driver.execute_script("arguments[0].click();", btn)
                        print("  ✅ 팝업 닫기")
                        time.sleep(1)
            except Exception:
                continue
    except Exception:
        pass

def accept_cookies(driver):
    try:
        time.sleep(3)
        cookie_selectors = [
            "#truste-consent-button",
            "button[id*='accept']", "button[class*='accept']", ".truste-button",
        ]
        for selector in cookie_selectors:
            try:
                for btn in driver.find_elements(By.CSS_SELECTOR, selector):
                    if btn.is_displayed():
                        driver.execute_script("arguments[0].click();", btn)
                        print("  ✅ 쿠키 동의 완료")
                        time.sleep(2)
                        return True
            except Exception:
                continue

        for text in ['Accept All','Accept','Accetta','Continua','동의','수락','Agree']:
            try:
                els = driver.find_elements(By.XPATH, f"//*[contains(., '{text}')]")
                for el in els:
                    if el.is_displayed() and el.tag_name in ['button','a']:
                        driver.execute_script("arguments[0].click();", el)
                        print(f"  ✅ '{text}' 클릭 완료")
                        time.sleep(2)
                        return True
            except Exception:
                continue
        print("  ⚠️ 쿠키 버튼 없음")
        return False
    except Exception as e:
        print(f"  ⚠️ 쿠키 처리 에러: {e}")
        return False

def is_hq_path(url: str) -> bool:
    p = urlparse(url)
    host = p.netloc.lower().replace('www.', '')
    parts = [x for x in p.path.split('/') if x]
    return host.endswith(TARGET_DOMAIN) and len(parts) >= 1 and parts[0] == HQ_SITE_CODE

def wait_dom_settled(driver, timeout=20):
    end = time.time() + timeout
    stable = 0
    last = -1
    while time.time() < end:
        h = driver.execute_script("return document.body ? document.body.scrollHeight : 0")
        if abs(h - last) < 10:
            stable += 1
            if stable >= 3:
                break
        else:
            stable = 0
        last = h
        time.sleep(0.8)

def wait_key_elements(driver, timeout=15):
    sels = [".cp-hero-kv", ".kv-wrpr", "#gnb", "header", "footer"]
    try:
        WebDriverWait(driver, timeout).until(
            lambda d: any(len(d.find_elements(By.CSS_SELECTOR, s))>0 for s in sels)
        )
    except Exception:
        pass

# =========================
# 에러 페이지 감지
# =========================
# 브랜드 에러 페이지: <title>error | Example Site</title> 등
# HTTP 에러: <title>502 Bad Gateway</title> 등
_ERROR_TITLE_KEYWORDS = ['error', '404', '502', '503', 'bad gateway', 'page not found', 'not available']

def is_error_page(driver) -> bool:
    """에러/없는 페이지 여부 판단 (title + canonical URL + HQ 전용 요소)"""
    # 1. 영어 title 키워드 (기존)
    try:
        title = driver.title.lower().strip()
        if any(kw in title for kw in _ERROR_TITLE_KEYWORDS):
            return True
    except Exception:
        pass

    # 2. 브랜드 공통 에러 페이지: canonical URL에 /common/error/ 또는 /common/404/ 포함
    try:
        canonical = driver.execute_script(
            "var el=document.querySelector('link[rel=\"canonical\"]'); return el?el.href:'';"
        )
        if canonical and any(p in canonical for p in ('/common/error/', '/common/404/')):
            return True
    except Exception:
        pass

    # 3. HQ(한국) 전용 에러 구조: aiscPrivateError 요소가 "실제로 표시"된 경우만
    #    ⚠ 정상 HQ 페이지도 이 에러 컨테이너를 hidden 으로 DOM 에 품고 있어,
    #      find_elements(존재 여부)만 보면 정상 페이지가 오탐되어 skip 된다.
    #      → is_displayed() 로 화면에 실제 노출된 경우로 한정.
    try:
        if any(e.is_displayed() for e in driver.find_elements(By.ID, 'aiscPrivateError')):
            return True
    except Exception:
        pass

    # 4. Chrome 브라우저 에러 페이지 (ERR_TOO_MANY_REDIRECTS 등 네트워크 에러)
    #    ⚠ 정상 페이지 본문/스크립트에 ERR_ 문자열이 우연히 있어도 오탐하지 않도록,
    #      Chrome 에러 인터스티셜(=title 이 비어있음)인 경우로 한정.
    try:
        src = driver.page_source
        if src and not (driver.title or '').strip() and any(e in src for e in ('ERR_TOO_MANY_REDIRECTS', 'ERR_CONNECTION', 'ERR_NAME_NOT_RESOLVED', 'ERR_TIMED_OUT')):
            return True
    except Exception:
        pass

    return False

def get_main_document_status(driver, final_url):
    """perf log에서 메인 문서(Document) 응답의 HTTP status 반환. 못 찾으면 None.
    Selenium 은 status 를 직접 안 줘서, chromedriver 의 performance log 로
    Network.responseReceived 이벤트를 파싱해 메인 프레임 응답 코드를 얻는다."""
    def _n(u): return (u or "").split('#')[0].rstrip('/')
    target = _n(final_url)
    status = None
    try:
        for entry in driver.get_log("performance"):
            try:
                msg = json.loads(entry["message"])["message"]
            except Exception:
                continue
            if msg.get("method") != "Network.responseReceived":
                continue
            p = msg.get("params", {})
            if p.get("type") != "Document":           # iframe/XHR 제외, 메인 문서만
                continue
            resp = p.get("response", {})
            if _n(resp.get("url")) == target:          # 최종 문서 url 매칭
                status = resp.get("status")            # 여러 개면 마지막(최종) 사용
    except Exception:
        pass
    return status

def is_unknown_page(driver, url) -> bool:
    """존재하지 않는 캠페인 경로에 대해 메인 도메인이 HTTP 200 + 홈 fallback 을
    돌려주는 "알 수 없는 페이지" 감지. 정상 마케팅 페이지는 <meta property="og:url"> 이
    실제 URL 로 채워지지만, unknown 페이지는 그 값이 비어있거나(EMPTY) 태그가 아예 없다(MISSING).
    → 리다이렉트/HTTP status/기존 error 마커로는 안 잡히는(200) 케이스를 여기서 걸러 skip.
    (기존 캡처 대조: HQ 10 + 기타 site 22 = 32개 정확히 감지, 정상 87 오탐 0)
    """
    try:
        host = urlparse(url).netloc.lower()
    except Exception:
        return False
    # 검사 대상은 메인 도메인 루트/www 만 (shop.* / example-partner / *.com.cn 등은
    # 원래 og:url 이 없을 수 있어 제외 — 오탐 방지)
    if host not in OG_CHECK_HOSTS:
        return False
    try:
        og = driver.execute_script(
            "var el=document.querySelector('meta[property=\"og:url\"]');"
            "return el ? (el.getAttribute('content') || '') : '__MISSING__';"
        )
    except Exception:
        return False
    return og == '__MISSING__' or (og or '').strip() == ''

# =========================
# 스크린샷 / 스크롤
# =========================
def screenshot_png(driver): return driver.get_screenshot_as_png()

def cdp_with_timeout(driver, cmd, params=None, timeout=None):
    """CDP 명령을 별도 스레드에서 실행하고 제한 시간을 넘기면 TimeoutError 를 낸다.

    ⚠ set_page_load_timeout / set_script_timeout 은 CDP 에 적용되지 않는다.
      Page.captureSnapshot 이 무응답이면 워커가 영구 대기하고, 그 future 하나 때문에
      as_completed 루프가 끝나지 않아 결과 CSV 기록과 드라이버 정리까지 통째로 막힌다.
      (2026-07-21 실제 사고) 스레드는 daemon 이라 끊어도 프로세스 종료를 막지 않는다.
    """
    timeout = CDP_TIMEOUT if timeout is None else timeout
    box = {}
    def _call():
        try: box["r"] = driver.execute_cdp_cmd(cmd, params or {})
        except Exception as e: box["e"] = e
    th = threading.Thread(target=_call, daemon=True)
    th.start(); th.join(timeout)
    if th.is_alive():
        raise TimeoutError(f"CDP {cmd} 응답 없음 ({timeout}초 초과)")
    if "e" in box: raise box["e"]
    return box.get("r")

def looks_blank(png_bytes):
    img = Image.open(io.BytesIO(png_bytes)).convert("L")
    arr = np.array(img)
    if arr.size == 0: return True
    nonwhite = np.sum(arr < 250)
    return (nonwhite / arr.size) < 0.005

# position:fixed / sticky 를 static 으로 바꾼다(원복용으로 바뀐 요소를 window 에 보관).
# ⚠ 컷마다 다시 호출되므로 목록을 덮어쓰지 말고 누적해야 한다.
#   덮어쓰면(두 번째 호출은 대개 빈 배열) 원복이 안 돼 position:static 이 주입된 채로
#   MHTML 이 저장된다 — 아카이브 원본이 훼손되므로 반드시 누적 + 전량 원복.
_JS_STICKY_OFF = """
window.__pcSticky = window.__pcSticky || [];
var n = 0;
document.querySelectorAll('*').forEach(function (el) {
  var p = getComputedStyle(el).position;
  if (p === 'fixed' || p === 'sticky') {
    window.__pcSticky.push([el, el.style.getPropertyValue('position'),
                                el.style.getPropertyPriority('position')]);
    el.style.setProperty('position', 'static', 'important');
    n++;
  }
});
return n;
"""
_JS_STICKY_ON = """
var c = window.__pcSticky || [];
for (var i = c.length - 1; i >= 0; i--) {          // 나중에 바꾼 것부터 되돌린다
  var el = c[i][0], v = c[i][1], pr = c[i][2];
  if (v) { el.style.setProperty('position', v, pr || ''); }
  else   { el.style.removeProperty('position'); }   // 원래 인라인 스타일이 없었으면 지운다
}
window.__pcSticky = null;
return c.length;
"""

def capture_full_page_mobile(driver, width):
    """모바일 전체페이지를 스크롤하며 찍어 이어붙인다.

    ⚠ 좌표 단위 주의 — 예전 버전의 버그 원인:
      스크롤 값(scrollTo/innerHeight)은 CSS px 인데 스크린샷은 device px(=CSS × 배율 3)다.
      옛 코드는 겹침 보정에 CSS px 값(100)을 device px 자리에 그대로 써서
        ① 이음새마다 200 device px 씩 내용이 중복되고
        ② 캔버스가 실제보다 100×(장수-1) px 커져 그만큼 하단이 검정으로 남았다.
        (실측: 높이 75,492px = 2432×31+100, 하단 검정 3,000px = 100×30 으로 정확히 일치)
      → 이제 "각 컷을 실제 스크롤 위치 × 배율" 자리에 그대로 붙인다. 겹침 보정 산술이 사라져
        단위 혼동 자체가 생길 수 없고, 캔버스 높이도 붙인 범위에서 역산하므로 빈칸이 안 생긴다.
    """
    driver.execute_script("window.scrollTo(0,0);")
    time.sleep(2)

    sticky_n = 0
    if NEUTRALIZE_STICKY:
        try:
            sticky_n = driver.execute_script(_JS_STICKY_OFF) or 0
            if sticky_n:
                print(f"  📌 sticky/fixed {sticky_n}개 → static (이음새 반복 방지)")
            time.sleep(0.5)
        except Exception:
            pass

    def _collect(overlap):
        """겹침(CSS px)을 정해 스크롤하며 컷을 모은다. → [(이미지, 실제 스크롤위치[CSS px])]"""
        total_height = driver.execute_script(
            "return Math.max(document.body.scrollHeight, document.documentElement.scrollHeight);")
        vh = driver.execute_script("return window.innerHeight")
        shots, offset = [], 0
        while True:
            driver.execute_script(f"window.scrollTo(0,{offset});")
            time.sleep(2)
            if NEUTRALIZE_STICKY:
                # JS 로 매 스크롤마다 위치를 다시 잡는 사이트가 있어 컷마다 다시 눌러준다
                try: driver.execute_script(_JS_STICKY_OFF)
                except Exception: pass
            # 요청한 offset 이 아니라 "실제로 스크롤된 위치"를 쓴다 —
            # 페이지 끝에서는 요청보다 덜 내려가므로, 이 값이 붙일 좌표의 진실이다.
            y = int(driver.execute_script("return Math.round(window.pageYOffset);") or 0)
            shots.append((Image.open(io.BytesIO(driver.get_screenshot_as_png())), y))
            if y + vh >= total_height - 1:
                break
            if len(shots) >= MOBILE_MAX_SHOTS:
                # 여기서 끊으면 페이지 아래쪽이 통째로 빠진다 — 조용히 넘어가지 않는다
                print(f"  ⚠️ 컷 {MOBILE_MAX_SHOTS}장 상한 도달 → 이 아래 "
                      f"약 {max(total_height - (y + vh), 0):,} CSS px 는 캡처되지 않음")
                break
            nxt = offset + vh - overlap
            if nxt <= offset:       # 뷰포트보다 겹침이 크면 무한루프 — 방어
                break
            offset = nxt
        return shots, vh

    def _sticky_band(shots):
        """컷들의 "상단 몇 행이 서로 똑같은지" 를 재서 스티키 바 높이(device px)를 구한다.

        CSS position 을 static 으로 바꿔도 JS 가 스크롤마다 위치를 다시 잡는 사이트가 있어
        (2026-07-21 사례: 101개를 static 으로 바꿨는데도 570px 바가 32회 반복)
        "왜 고정됐는지" 를 따지지 않고 결과 픽셀에서 직접 잰다.
        """
        if len(shots) < 3:
            return 0
        limit = shots[0][0].height // 3          # 뷰포트의 1/3 이상은 스티키로 보지 않음
        best = 0
        for i in range(1, len(shots) - 1):
            a1 = np.asarray(shots[i][0].convert("RGB"))
            a2 = np.asarray(shots[i + 1][0].convert("RGB"))
            n = 0
            while n < limit and np.array_equal(a1[n], a2[n]):
                n += 1
            best = max(best, n)                  # 바가 숨는 구간이 있으므로 최대값을 쓴다
        return best

    try:
        overlap = MOBILE_STITCH_OVERLAP
        shots, vh = _collect(overlap)
        fw = shots[0][0].width
        scale = (fw / float(width)) if width else 1.0     # device px / CSS px (보통 3.0)
        band = _sticky_band(shots)

        # ⚠ 잘라낸 밴드 자리는 "앞 컷"이 덮어줘야 빈칸이 안 생긴다.
        #   앞 컷이 뒤 컷 시작점 너머로 뻗는 길이 = 겹침 × 배율 이므로, 겹침이 밴드보다 커야 한다.
        #   부족하면 겹침을 늘려 한 번만 다시 찍는다.
        if band and band > overlap * scale and len(shots) > 1:
            need = min(int(band / scale) + 20, max(int(vh * 0.5), MOBILE_STITCH_OVERLAP))
            if need > overlap:
                print(f"  🔁 스티키 바 {band}px 감지 → 겹침 {overlap}→{need}px 로 재촬영")
                overlap = need
                shots, vh = _collect(overlap)
                band = _sticky_band(shots)
        if band:
            print(f"  ✂️  스티키 바 {band}px — 첫 컷 제외하고 잘라냄 (이음새 반복 제거)")
    finally:
        # ⚠ 조건 없이 원복한다 — 첫 호출이 0개여도 컷마다 재적용하며 더 잡았을 수 있고,
        #   여기서 원복해야 뒤이어 저장되는 MHTML 이 원본 그대로가 된다.
        if NEUTRALIZE_STICKY:
            try:
                restored = driver.execute_script(_JS_STICKY_ON) or 0
                if restored: print(f"  ↩️  sticky/fixed {restored}개 원복 (MHTML 은 원본 상태로 저장)")
            except Exception: pass

    if len(shots) == 1:
        return shots[0][0]

    def _top(y): return int(round(y * scale))
    fh = max(_top(y) + img.height for img, y in shots)
    final = Image.new('RGB', (fw, fh), (255, 255, 255))   # 빈칸이 남아도 검정이 아닌 흰색
    for i, (img, y) in enumerate(shots):
        if i and band:                       # 첫 컷의 상단은 진짜 페이지 머리라 남긴다
            img = img.crop((0, band, fw, img.height))
            final.paste(img, (0, _top(y) + band))
        else:
            final.paste(img, (0, _top(y)))   # 겹치는 부분은 뒤 컷이 덮어씀(동일 내용)

    # ── 하단 여백 트림 ────────────────────────────────────────
    # 마지막 컷의 뷰포트가 페이지 끝을 넘어서면 그만큼 흰 띠가 남는다
    # (2026-07-22 실측: 한 건에서 2,161px). 콘텐츠가 끝난 지점까지만 잘라낸다.
    if TRIM_TRAILING_BLANK:
        arr = np.asarray(final.convert("L"))
        nonblank = np.where(arr.min(axis=1) < 245)[0]     # 흰색이 아닌 행
        if nonblank.size:
            end = int(nonblank[-1]) + 1 + TRIM_KEEP_MARGIN
            if end < final.height:
                cut = final.height - end
                final = final.crop((0, 0, fw, min(end, final.height)))
                print(f"  ✂️  하단 여백 {cut:,}px 잘라냄")
    return final

def smooth_scroll_desktop(driver):
    last = driver.execute_script("return document.body.scrollHeight")
    for _ in range(10):
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(2)
        new = driver.execute_script("return document.body.scrollHeight")
        if new == last: break
        last = new
    return last

# =========================
# Chrome 드라이버 (워커별 재사용)
# =========================
# 태스크는 device 순으로 정렬해 넣으므로, 한 스레드는 대개 같은 device 를 연속 처리한다.
# device 가 바뀌는 순간에만 기존 driver 를 닫고 새로 만든다(= 동시 Chrome 수는 워커 수와 같음).
_tls = threading.local()
_all_drivers = []                 # 종료 시 일괄 정리용
_all_drivers_lock = threading.Lock()

def _build_options(device_type):
    o = Options()
    o.add_argument('--headless=new')
    o.add_argument('--disable-gpu')
    o.add_argument('--no-sandbox')
    o.add_argument('--disable-dev-shm-usage')
    o.add_argument('--hide-scrollbars')
    # perf log(goog:loggingPrefs) 활성화 후 --headless=new 가 빈 창을 띄우는 현상 방지 → 창을 화면 밖으로
    o.add_argument('--window-position=-32000,-32000')
    if device_type == "MO":
        o.add_argument(f'--user-agent={MOBILE_UA}')
        vw, vh = MOBILE_VIEWPORT
    else:
        vw, vh = DESKTOP_VIEWPORT
    o.add_argument(f'--window-size={vw},{vh}')
    # perf log 활성화 → Network.responseReceived 로 메인 문서 HTTP status 획득 (404 감지용)
    o.set_capability("goog:loggingPrefs", {"performance": "ALL"})
    return o, vw, vh

def _new_driver(device_type):
    o, vw, vh = _build_options(device_type)
    d = webdriver.Chrome(options=o)
    d.set_page_load_timeout(PAGE_LOAD_TIMEOUT)
    d.set_script_timeout(SCRIPT_TIMEOUT)
    if device_type == "MO":
        d.execute_cdp_cmd('Emulation.setDeviceMetricsOverride', {
            'width': vw, 'height': vh, 'deviceScaleFactor': MOBILE_SCALE, 'mobile': True
        })
    with _all_drivers_lock:
        _all_drivers.append(d)
    return d, vw, vh

def get_driver(device_type):
    """이 스레드가 쓸 driver 를 돌려준다. REUSE_DRIVER 면 재사용, 아니면 매번 새로."""
    if not REUSE_DRIVER:
        return _new_driver(device_type)
    cur = getattr(_tls, "driver", None)
    if cur is not None and getattr(_tls, "device", None) == device_type:
        try:
            # 세션이 살아있는지 확인 (죽었으면 예외 → 새로 만든다)
            _ = cur.current_url
            # ── URL 간 상태 격리 ──
            # 쿠키를 남기면 쿠키배너 클릭 여부·로그인 상태가 다음 URL 판정을 바꾼다.
            # perf log 도 비워야 다음 URL 의 HTTP status 를 잘못 읽지 않는다.
            try: cur.delete_all_cookies()
            except Exception: pass
            try: cur.get_log("performance")   # 읽어서 비움
            except Exception: pass
            return cur, _tls.vw, _tls.vh
        except Exception:
            _quit_driver(cur)
    elif cur is not None:
        _quit_driver(cur)                     # device 가 바뀌면 닫고 새로
    d, vw, vh = _new_driver(device_type)
    _tls.driver, _tls.device, _tls.vw, _tls.vh = d, device_type, vw, vh
    return d, vw, vh

def _quit_driver(d, timeout=None):
    """driver 를 닫는다. 응답이 없으면 프로세스를 직접 죽인다.

    ⚠ 이게 hang 의 진짜 원인이었다 (2026-07-22 스택 덤프로 확인).
      driver.quit() → service.stop() → send_remote_shutdown_command() → is_connectable() 은
      urllib 로 chromedriver 에 요청을 거는데 **타임아웃이 없다**. chromedriver 가 응답을 멈추면
      이 호출이 영구히 블록되고, 그 스레드의 태스크는 결과 행조차 남기지 못한 채 사라진다.
      (개별 명령 타임아웃·CDP 타임아웃은 이 경로를 전혀 감시하지 못한다)
    """
    if d is None:
        return
    timeout = QUIT_TIMEOUT if timeout is None else timeout

    def _q():
        try:
            d.quit()
        except Exception:
            pass

    th = threading.Thread(target=_q, daemon=True)
    th.start()
    th.join(timeout)
    if th.is_alive():
        print(f"  ⚠️ driver.quit() 무응답({timeout}초) → chromedriver 프로세스 강제 종료")
        try:
            d.service.process.kill()          # 종료 요청을 기다리지 않고 끊는다
        except Exception:
            pass
    with _all_drivers_lock:
        if d in _all_drivers:
            _all_drivers.remove(d)

def drop_thread_driver():
    """현재 스레드의 driver 를 버린다 (세션이 깨졌을 때 다음 시도에서 새로 만들도록)."""
    _quit_driver(getattr(_tls, "driver", None))
    _tls.driver = None

def quit_all_drivers():
    with _all_drivers_lock:
        drivers = list(_all_drivers)
    for d in drivers:
        _quit_driver(d)

# =========================
# CDP 전체 페이지 캡처
# =========================
def capture_fullpage_cdp(driver):
    """CDP 로 전체 페이지를 한 번에 캡처해 PNG bytes 반환. 실패하면 None.
    스크롤 스티칭(조각 합성)·창 크기 늘리기를 모두 대체한다."""
    try:
        m = driver.execute_cdp_cmd("Page.getLayoutMetrics", {})
        size = m.get("cssContentSize") or m.get("contentSize") or {}
        w = int(size.get("width") or 0)
        h = int(size.get("height") or 0)
        if w <= 0 or h <= 0:
            return None
        if h > MAX_CAPTURE_HEIGHT:
            print(f"  ⚠️ 페이지가 너무 김({h}px) → {MAX_CAPTURE_HEIGHT}px 로 잘라 캡처")
            h = MAX_CAPTURE_HEIGHT
        res = cdp_with_timeout(driver, "Page.captureScreenshot", {
            "format": "png",
            "captureBeyondViewport": True,
            "clip": {"x": 0, "y": 0, "width": w, "height": h, "scale": 1},
        })
        data = res.get("data") or ""
        return base64.b64decode(data) if data else None
    except Exception as e:
        print(f"  ⚠️ CDP 캡처 실패 → 기존 방식으로 대체: {e}")
        return None

def build_output_paths(url, device_type, timestamp):
    """이 (url, device) 가 저장될 PNG/MHTML 경로. 브라우저 없이 계산 가능해야
    resume(이미 있으면 skip) 판정을 기동 전에 할 수 있다."""
    sitecode, page_path, _ = get_page_info(url)
    parsed = urlparse(url)
    # 쿼리 파라미터 있는 경우 파일명에 포함
    # (ex: ES_PC_offer_black-friday_page_category-tv-audio_1112.png / 없으면 ES_PC_offer_black-friday_page_1112.png)
    query_part = parsed.query.replace('=', '-').replace('&', '_') if parsed.query else ''
    if query_part:
        base = safe_filename(f'{sitecode}_{device_type}_{page_path}_page_{query_part}_{timestamp}.png')
    else:
        base = safe_filename(f'{sitecode}_{device_type}_{page_path}_page_{timestamp}.png')
    png = f"{OUTPUT_DIR}/{base}"
    return png, png[:-4] + ".mhtml"


def already_captured(png_path, mhtml_path):
    """이 (url, device) 를 오늘 이미 찍었는지 — OUTPUT_DIR 루트 + sitecode 폴더 양쪽을 본다.

    ⚠ 루트만 보면 안 된다: 캡처 후 foldering_move_png.py 로 파일을 {OUTPUT_DIR}/{SITECODE}/
      아래로 옮기고 나면 루트가 비어, 재실행이 이미 찍은 것까지 전부 다시 찍는다
      (2026-07-23: 결번 몇 건만 보충하려던 재실행이 전량 재캡처가 될 뻔했다).
      폴더명 규칙은 foldering 스크립트와 동일 — 파일명에서 _PC / _MO 앞부분, 공백은 언더바.
    """
    if os.path.exists(png_path) and os.path.exists(mhtml_path):
        return True
    base = os.path.basename(png_path)
    m = re.match(r"(.+?)[ _](PC|MO)", base)
    if not m:
        return False
    sub = os.path.join(OUTPUT_DIR, m.group(1).replace(" ", "_"))
    return (os.path.exists(os.path.join(sub, base))
            and os.path.exists(os.path.join(sub, os.path.basename(mhtml_path))))


def wait_page_settled(driver, timeout=None):
    """readyState 가 complete 될 때까지만 기다린다 (고정 sleep 대체)."""
    timeout = PAGE_SETTLE_TIMEOUT if timeout is None else timeout
    try:
        WebDriverWait(driver, timeout).until(
            lambda d: d.execute_script("return document.readyState") == "complete"
        )
    except Exception:
        pass

# =========================
# 메인 캡처 함수
# =========================
def capture_page(url, device_type):
    """캡처 1건 수행. 결과를 dict 로 반환 (result_*.csv 한 행에 대응).
    result: ok / redirected / error_page / unknown_page / timeout / error"""
    sitecode, page_path, sitetype = get_page_info(url)
    page_name = safe_filename(f"{sitecode}_{page_path}")
    timestamp = datetime.now().strftime("%m%d")

    # ── 결과 기록용 — 어느 디바이스가 왜 skip 됐는지 사후 검수 가능하게 남긴다 ──
    started = time.time()
    info = {"url": url, "device": device_type, "sitecode": sitecode, "result": "",
            "final_url": "", "http_status": "", "title": "", "detail": "", "elapsed": 0.0}

    def _ret(result, detail=""):
        info["result"] = result
        info["detail"] = detail
        info["elapsed"] = round(time.time() - started, 1)
        return info

    # ── 이어하기(resume) — 이미 받은 건 브라우저를 띄우지도 않는다 ──
    png_path, mhtml_path = build_output_paths(url, device_type, timestamp)
    if SKIP_IF_EXISTS and already_captured(png_path, mhtml_path):
        print(f"  ⏩ 이미 있음 → skip: {os.path.basename(png_path)}")
        return _ret("exists", "PNG/MHTML 이미 존재")

    # ⚠ Chrome 확보도 try 안에서 — 기동 자체가 실패해도 결과 집계를 타게 한다
    #   (밖에 두면 예외가 capture_page 밖으로 튀어 그 작업이 결과에서 통째로 사라짐)
    driver = None
    try:
        driver, vw, vh = get_driver(device_type)
        _register_driver_for_deadline(driver)   # 상한 초과 시 감시 스레드가 이 driver 를 끊는다
        print(f"  📸 {device_type} 버전 캡처 중...")
        driver.get(url)
        wait_page_settled(driver)

        # ── 리다이렉트 감지 ──────────────────────────────────────
        final_url = driver.current_url
        info["final_url"] = final_url
        try:
            info["title"] = (driver.title or "").strip()
        except Exception:
            pass
        _norm = normalize_url_for_compare   # 추적파라미터·fragment·www 등 무시하고 비교
        def _login_hit(u):
            """로그인/인증 화면 URL 이면 걸린 키워드를 돌려준다 (아니면 None)"""
            if not RECHECK_URL_BEFORE_SAVE:
                return None
            return next((k for k in SKIP_URL_KEYWORDS if k in (u or "").lower()), None)

        # 로그인 화면은 리다이렉트보다 먼저 판정 — 같은 skip 이라도 사유를 구분해 따로 집계한다
        _hit = _login_hit(final_url)
        if _hit:
            print(f"  🔒 로그인/인증 화면 감지 → skip")
            print(f"      이동: {final_url}")
            return _ret("login_page", f"URL 에 '{_hit}' 포함")
        if _norm(final_url) != _norm(url):
            print(f"  ⏭️  리다이렉트 감지 → skip")
            print(f"      원본: {url}")
            print(f"      이동: {final_url}")
            return _ret("redirected", "final_url != url")

        # ── HTTP 상태코드 감지 (404/5xx) — 언어·마크업 무관 ──────
        # Selenium 은 status 를 직접 안 줘서 perf log 로 메인 문서 응답 status 확인.
        # HQ 처럼 한국어 error 페이지라 is_error_page 가 못 잡는 진짜 404 를 잡는다.
        status = get_main_document_status(driver, final_url)
        info["http_status"] = "" if status is None else status
        if status is not None and status >= 400:
            print(f"  ⛔ HTTP {status} 감지 → skip")
            return _ret("error_page", f"HTTP {status}")

        # ── 에러 페이지 감지 (404 / 502 / 브랜드 error 등) ─────
        if is_error_page(driver):
            print(f"  ⛔ 에러 페이지 감지 (title: '{driver.title}') → skip")
            return _ret("error_page", "is_error_page")

        # ── unknown(soft-404) 페이지 감지 (HTTP 200 인데 og:url 없음 → 홈 fallback) ─
        # 존재하지 않는 캠페인 경로가 200 을 주고 홈 컨텐츠를 렌더해 위 감지들을 다 빠져나가는 케이스.
        if is_unknown_page(driver, final_url):
            print(f"  ⛔ unknown 페이지 감지 (og:url 비었음/없음, title: '{driver.title}') → skip")
            return _ret("unknown_page", "og:url EMPTY/MISSING")

        close_popups(driver)
        accept_cookies(driver)
        time.sleep(2)
        close_popups(driver)

        # ── 저장 직전 URL 재확인 (지연 리다이렉트 / 로그인 화면 차단) ──────
        # 위 리다이렉트 판정은 get() 직후 1회뿐이라, 쿠키·팝업 처리 중 JS 로 늦게 이동하는
        # 로그인 페이지가 그대로 저장된다(캡처된 mhtml 의 Snapshot-Content-Location 으로 확인 가능).
        if RECHECK_URL_BEFORE_SAVE:
            cur = driver.current_url
            hit = _login_hit(cur)
            if hit:
                info["final_url"] = cur
                print(f"  🔒 로그인/인증 화면 감지(저장 직전) → skip")
                print(f"      이동: {cur}")
                return _ret("login_page", f"저장 직전 재확인 · URL 에 '{hit}' 포함")
            if _norm(cur) != _norm(url):
                info["final_url"] = cur
                print(f"  ⏭️  지연 리다이렉트 감지(저장 직전) → skip")
                print(f"      이동: {cur}")
                return _ret("redirected", "저장 직전 재확인에서 이동 감지")

        filename, mhtml_filename = png_path, mhtml_path

        if is_hq_path(url):
            wait_dom_settled(driver)
            wait_key_elements(driver)

        png = None
        if USE_CDP_FULLPAGE:
            # 레이지 로딩 이미지를 띄우려면 한 번은 끝까지 훑어야 한다
            smooth_scroll_desktop(driver)
            driver.execute_script("window.scrollTo(0,0)")
            png = capture_fullpage_cdp(driver)

        if png is None:
            # ── 기본 경로: 기존 방식 (PC=창 늘리기 / MO=스크롤 스티칭) ──
            if device_type == "MO":
                # capture_full_page_mobile 이 스스로 끝까지 스크롤하며 찍으므로
                # 여기서 smooth_scroll_desktop 을 또 돌리면 시간만 두 배로 든다.
                img = capture_full_page_mobile(driver, vw)
                img.save(filename, 'PNG', optimize=True)
            else:
                total_h = smooth_scroll_desktop(driver)
                driver.set_window_size(vw, min(total_h + 200, MAX_CAPTURE_HEIGHT))
                time.sleep(2)
                png = screenshot_png(driver)

        if png is not None:
            # blank 재시도는 PC/MO 공통 적용 (예전엔 PC 에만 있었다)
            retry = 2
            while looks_blank(png) and retry > 0:
                print("  ⚠️ 화면이 비어 보여 재시도...")
                time.sleep(4)
                wait_dom_settled(driver)
                png2 = capture_fullpage_cdp(driver) if USE_CDP_FULLPAGE else None
                png = png2 if png2 is not None else screenshot_png(driver)
                retry -= 1
            with open(filename, "wb") as f:
                f.write(png)
        print(f"  ✅ 저장: {filename}")

        # ── MHTML 저장 ────────────────────────────────────────
        # fallback 으로 창을 늘렸다면 원래 뷰포트로 복원 (그 상태로 CDP 호출 시 blank 발생)
        if device_type != "MO":
            driver.set_window_size(vw, vh)
        driver.execute_script("window.scrollTo(0,0)")
        try:
            result = cdp_with_timeout(driver, "Page.captureSnapshot", {"format": "mhtml"})
        except TimeoutError as e:
            # 이 세션은 CDP 가 응답을 멈춘 상태 — 물고 다음 URL 로 가면 그 URL 까지 망친다.
            print(f"  ⚠️ MHTML 저장 무응답 → 이 건만 포기: {e}")
            drop_thread_driver()
            return _ret("mhtml_failed", str(e))
        mhtml_data = (result or {}).get("data", "")
        if not mhtml_data or len(mhtml_data) < MIN_MHTML_BYTES:
            # PNG 는 남았는데 MHTML 만 실패한 경우 — 결과와 산출물이 어긋나지 않게 따로 표시
            print(f"  ⚠️ MHTML 이 비었거나 너무 작음({len(mhtml_data)}B)")
            return _ret("mhtml_failed", f"mhtml {len(mhtml_data)}B < {MIN_MHTML_BYTES}B")
        with open(mhtml_filename, "wb") as mf:  # bytes 쓰기 (텍스트 모드 시 blank 가능)
            mf.write(mhtml_data.encode("utf-8"))
        print(f"  ✅ MHTML 저장: {mhtml_filename}")
        return _ret("ok")

    except TimeoutException as e:
        # PAGE_LOAD_TIMEOUT / SCRIPT_TIMEOUT 초과 — 실패로 확정하고 다음 작업으로
        print(f"  ⏱️  타임아웃({PAGE_LOAD_TIMEOUT}초) → skip: {url}")
        # 재사용 driver 는 로딩이 걸린 채로 남아 다음 URL 까지 망칠 수 있으니 로딩을 끊는다
        try:
            driver.execute_script("window.stop();")
        except Exception:
            drop_thread_driver()
        return _ret("timeout", str(e).split('\n')[0][:200])
    except Exception as e:
        print(f"❌ 에러: {e}")
        import traceback; traceback.print_exc()
        # 재사용 중인 driver 가 깨졌을 수 있으니 버린다 (다음 작업은 새 Chrome 으로)
        drop_thread_driver()
        return _ret("error", f"{type(e).__name__}: {e}".split('\n')[0][:200])
    finally:
        # 재사용 모드에선 여기서 닫지 않는다 (전체 종료 시 quit_all_drivers 로 일괄 정리)
        if driver is not None and not REUSE_DRIVER:
            _quit_driver(driver)

# =========================
# 일일 리포트 (최신이 항상 왼쪽)
# =========================
def write_daily_report(csv_path, run_date, quiet=False):
    """날짜별 결과를 한 파일에 누적한다 — 최신 날짜가 B~D열, 과거는 오른쪽으로 밀린다.

    구조:  A열 = url,  이후 하루당 3열 = [{날짜} PC, {날짜} MO, {날짜} 이슈]
    같은 날 다시 돌리면 그 날짜 블록을 덮어쓴다(열이 중복으로 늘지 않는다).
    URL 은 어제까지 있던 것도 계속 남긴다 — "어제는 있었는데 오늘 아예 안 돈" 것을 보이게 하려고.

    ⚠ 입력이 메모리의 rows 가 아니라 **디스크의 result CSV** 인 것이 핵심이다(2026-07-23).
      하드 크래시로 프로세스가 죽으면 finally·atexit 이 안 돌아 메모리의 결과는 통째로 사라진다.
      캡처 건별로 append 되는 CSV 를 원본으로 삼아야 중간에 죽어도 그 시점까지가 리포트에 남는다.
      실행 중에도 REPORT_FLUSH_EVERY 건마다 부른다.
    """
    if not DAILY_REPORT:
        return None
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except Exception as e:
        print(f"  ⚠️ 일일 리포트 건너뜀(openpyxl 없음): {e}")
        return None

    path = f"{OUTPUT_DIR}/{DAILY_REPORT_NAME}"
    N = DAILY_REPORT_COLS

    # ── 오늘 값 정리: url -> (PC결과, MO결과, 이슈)
    # CSV 는 append 방식이라 같은 (url, device) 가 두 번 있을 수 있다(재시도 등) → 뒤엣것이 최종.
    try:
        with open(csv_path, encoding="utf-8-sig", newline="") as cf:
            rows = list(csv.DictReader(cf))
    except FileNotFoundError:
        print(f"  ⚠️ 일일 리포트 건너뜀(결과 CSV 없음): {csv_path}")
        return None
    today = {}
    for r in rows:
        u = r.get("url") or ""
        if not u:
            continue
        today.setdefault(u, {})[r.get("device", "")] = r
    day_vals = {}
    for u, devs in today.items():
        pc = devs.get("PC", {}).get("result", "")
        mo = devs.get("MO", {}).get("result", "")
        issues = []
        for dev in ("PC", "MO"):
            r = devs.get(dev)
            if r and r.get("result") not in ("ok", "exists", "error_page"):
                issues.append(f"{dev}:{r.get('result')}"
                              + (f"({r.get('detail','')[:40]})" if r.get("detail") else ""))
        # 한쪽만 저장된 경우는 눈에 띄게 표시 — 오늘 PT_MO 같은 케이스가 여기 걸린다
        if pc == "ok" and mo not in ("ok", "exists"):
            issues.append("MO 미저장")
        if mo == "ok" and pc not in ("ok", "exists"):
            issues.append("PC 미저장")
        day_vals[u] = (pc, mo, " / ".join(issues))

    header = run_date          # 예: "0722"
    if os.path.exists(path):
        wb = load_workbook(path)
        ws = wb.active
        # 이미 같은 날짜 블록이 있으면 그 자리에 덮어쓴다
        existing = {ws.cell(row=1, column=c).value: c
                    for c in range(2, ws.max_column + 1, N)}
        if header in existing:
            base = existing[header]
        else:
            ws.insert_cols(2, N)      # 최신을 항상 B~D 로 — 과거 블록은 오른쪽으로 밀린다
            base = 2
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "daily"
        ws.cell(row=1, column=1, value="url")
        ws.cell(row=2, column=1, value="url")   # 2행은 소제목 행
        base = 2
        ws.insert_cols(2, N)

    # ── 헤더 2줄: 1행 = 날짜(병합), 2행 = PC / MO / 이슈
    ws.cell(row=1, column=1, value="url")
    ws.cell(row=1, column=base, value=header)
    ws.merge_cells(start_row=1, start_column=base, end_row=1, end_column=base + N - 1)
    for off, name in enumerate(("PC", "MO", "이슈")):
        ws.cell(row=2, column=base + off, value=name)

    # ── URL 행 위치 (A열) — 없으면 아래에 추가
    row_of = {}
    for r in range(3, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v:
            row_of[v] = r
    next_row = ws.max_row + 1 if ws.max_row >= 3 else 3
    for u in day_vals:
        if u not in row_of:
            ws.cell(row=next_row, column=1, value=u)
            row_of[u] = next_row
            next_row += 1

    # ── 값 채우기
    fill_bad = PatternFill("solid", fgColor="FFC7CE")     # 문제 있는 칸
    fill_ok = PatternFill("solid", fgColor="E2EFDA")      # 저장된 칸
    for u, r in row_of.items():
        pc, mo, issue = day_vals.get(u, ("(미실행)", "(미실행)", "오늘 대상 아님"))
        for off, val in enumerate((pc, mo, issue)):
            c = ws.cell(row=r, column=base + off, value=val)
            if off < 2:
                if val in ("ok", "exists"):
                    c.fill = fill_ok
                elif val not in ("error_page", ""):
                    c.fill = fill_bad
            elif val and val != "오늘 대상 아님":
                c.fill = fill_bad

    # ── 보기 설정
    ws.freeze_panes = "B3"
    ws.column_dimensions["A"].width = 72
    for off in range(N):
        col = get_column_letter(base + off)
        ws.column_dimensions[col].width = 14 if off < 2 else 40
    for cell in (ws.cell(row=1, column=1), ws.cell(row=1, column=base), ws.cell(row=2, column=1)):
        cell.font = Font(bold=True)
    ws.cell(row=1, column=base).alignment = Alignment(horizontal="center")
    for off in range(N):
        ws.cell(row=2, column=base + off).font = Font(bold=True)

    wb.save(path)
    if not quiet:
        n_issue = sum(1 for v in day_vals.values() if v[2])
        print(f"📊 일일 리포트 갱신: {path}  (최신 {header} = B~{get_column_letter(base + N - 1)}열, "
              f"이슈 {n_issue}건)")
    return path


# =========================
# HTTP 사전 필터
# =========================
def prefilter_dead_urls(urls):
    """브라우저를 띄우기 전에 상태코드만 확인해 "확실히 죽은" URL 을 걸러낸다.

    돌려주는 것 = {url: status}. 여기 담긴 URL 은 Chrome 을 태우지 않고 error_page 로 확정한다.

    ⚠ 판정을 넓히지 말 것 — 여기서 잘못 걸러내면 정상 페이지가 조용히 아카이브에서 빠진다.
      · 404/410/5xx 만 죽음으로 본다 (PREFILTER_DEAD_STATUS)
      · 403 은 봇 차단일 수 있어 제외 — 브라우저로 확인시킨다
      · 200 이지만 내용이 홈인 soft-404 는 여기서 못 잡는다 → 기존 og:url 검사가 계속 담당
      · 네트워크 오류·타임아웃도 제외 — 일시적일 수 있으니 브라우저에 맡긴다
    """
    if not PREFILTER_HTTP or not urls:
        return {}
    dead, t0 = {}, time.time()
    sess_headers = {"User-Agent": PREFILTER_UA, "Accept-Language": "en-US,en;q=0.9"}

    def _probe(u):
        try:
            # HEAD 를 안 받는 서버가 많아 GET + stream(본문 안 받음)으로 상태만 본다
            with requests.get(u, headers=sess_headers, timeout=PREFILTER_TIMEOUT,
                              allow_redirects=True, stream=True) as r:
                return u, r.status_code
        except Exception:
            return u, None          # 판단 보류 → 브라우저로 넘긴다

    print(f"🔎 HTTP 사전 확인 {len(urls)}개 (동시 {PREFILTER_WORKERS}) …")
    with ThreadPoolExecutor(max_workers=PREFILTER_WORKERS) as ex:
        for u, st in ex.map(_probe, urls):
            if st in PREFILTER_DEAD_STATUS:
                dead[u] = st
    print(f"   → 죽은 URL {len(dead)}개 확정 / 캡처 대상 {len(urls) - len(dead)}개 "
          f"({time.time() - t0:.0f}초)")
    return dead


# =========================
# 태스크 하드 데드라인 감시
# =========================
# 워커 스레드가 지금 무슨 작업을 언제 시작했는지 기록해두고, 감시 스레드가 초과분을 끊는다.
_inflight = {}                     # thread_ident -> {"url","device","t0","driver"}
_inflight_lock = threading.Lock()


def _mark_task_start(url, device):
    with _inflight_lock:
        _inflight[threading.get_ident()] = {"url": url, "device": device, "t0": time.time()}


def _mark_task_end():
    with _inflight_lock:
        _inflight.pop(threading.get_ident(), None)


def _register_driver_for_deadline(driver):
    """이 스레드가 지금 쓰는 driver 를 감시 대상에 붙인다(끊을 때 필요)."""
    with _inflight_lock:
        cur = _inflight.get(threading.get_ident())
        if cur is not None:
            cur["driver"] = driver


def start_deadline_watchdog(stop_event):
    """상한을 넘긴 태스크의 chromedriver 를 강제로 끊는 감시 스레드를 띄운다.

    driver 프로세스를 죽이면 그 워커의 Selenium 호출이 예외로 풀려나오므로,
    capture_page 의 except 가 받아 timeout/error 로 확정하고 다음 작업으로 넘어간다.
    (개별 명령 타임아웃으로는 "명령과 명령 사이"에서 늘어지는 hang 을 못 잡는다)
    """
    if not (TASK_DEADLINE_MO or TASK_DEADLINE_PC):
        return None

    def _loop():
        while not stop_event.wait(DEADLINE_CHECK_INTERVAL):
            now = time.time()
            with _inflight_lock:
                snapshot = list(_inflight.items())
            for ident, info in snapshot:
                limit = TASK_DEADLINE_MO if info["device"] == "MO" else TASK_DEADLINE_PC
                if not limit:
                    continue
                el = now - info["t0"]
                if el <= limit or info.get("killed"):
                    continue
                d = info.get("driver")
                print(f"  ⏰ 데드라인 초과 {el:.0f}초 > {limit}초 → 드라이버 강제 종료: "
                      f"{info['device']} {info['url']}")
                info["killed"] = True
                try:
                    # quit() 은 응답을 기다리므로 물린 상태에선 같이 멈춘다 → 프로세스를 직접 죽인다
                    d.service.process.kill()
                except Exception:
                    try:
                        d.quit()
                    except Exception:
                        pass

    th = threading.Thread(target=_loop, daemon=True, name="deadline-watchdog")
    th.start()
    return th


# =========================
# 여러 URL 병렬 캡처
# =========================
def capture_urls(urls, max_workers=MAX_WORKERS):
    # ⚠ 저장 폴더는 반드시 캡처 시작 "전"에 만든다.
    #   (예전엔 모든 캡처가 끝난 뒤에 만들어서, 폴더가 없으면 PNG/MHTML 저장이 전량 실패했다)
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # 로그 파일을 가장 먼저 연다 — 이후의 모든 출력(사전 확인·잔재 정리 포함)이 파일에 남는다
    ts = datetime.now().strftime("%m%d_%H%M")
    start_file_log(ts)

    if isinstance(urls,str):
        # strip 을 먼저 하고 '#' 판정 — 안 그러면 "  # 주석" 처럼 들여쓴 주석줄이 URL 로 들어간다
        _lines = [u.strip() for u in urls.split('\n')]
        urls = [u for u in _lines if u and not u.startswith('#')]

    # 이전 실행이 강제 종료되며 남긴 Chrome 을 먼저 치운다 (캡처 시작 전에만 안전)
    sweep_stale_chrome()

    # ── HTTP 사전 필터 ────────────────────────────────────────
    # 확실히 죽은 URL 은 Chrome 을 태우지 않고 여기서 확정한다(워커시간의 절반 이상이 이 부류였다).
    dead = prefilter_dead_urls(urls)
    live = [u for u in urls if u not in dead]

    # (url, device) 단위 태스크 — PC/MO 는 서로 독립이라 각각 별도 브라우저로 병렬 처리
    # ⚠ device 로 묶어서 넣는다: driver 재사용 시 한 스레드가 같은 device 를 연속 처리해야
    #   device 가 바뀔 때마다 Chrome 을 새로 띄우는 일이 없다.
    tasks = [(u, dev) for dev in ("PC", "MO") for u in live]
    print(f"\n🚀 총 {len(urls)}개 페이지 중 캡처 대상 {len(live)}개 × (PC/MO) = {len(tasks)}개 작업 "
          f"병렬 캡처 시작 (동시 {max_workers}개)\n")

    results = {}            # url -> {device: result 문자열}
    rows = []               # result_*.csv 용 (url, device) 단위 상세 행
    progress = {"n": 0}
    lock = threading.Lock()

    # ── 결과 CSV 를 캡처 1건마다 append ────────────────────────
    # 예전엔 모든 작업이 끝난 뒤 한 번에 썼다. 그래서 워커 하나가 물려 루프가 안 끝나면
    # 그날 판정 결과가 통째로 메모리에 갇힌 채 날아갔다(2026-07-21 사고).
    # 이제 진행 중에도 파일이 남아 중단돼도 어디까지 뭘 했는지 알 수 있다. (ts 는 위에서 만들었다)
    csv_path = f"{OUTPUT_DIR}/{RESULT_CSV_NAME}"
    _cols = ["url", "device", "sitecode", "result", "final_url", "http_status", "title", "detail", "elapsed"]
    csv_lock = threading.Lock()
    with open(csv_path, "w", encoding="utf-8-sig", newline="") as cf:
        csv.DictWriter(cf, fieldnames=_cols, extrasaction="ignore").writeheader()

    def _append_row(info):
        with csv_lock:
            with open(csv_path, "a", encoding="utf-8-sig", newline="") as cf:
                csv.DictWriter(cf, fieldnames=_cols, extrasaction="ignore").writerow(info)

    # 사전 필터로 확정된 죽은 URL 도 같은 CSV 에 남긴다 — 캡처를 안 했을 뿐 판정은 난 것이다.
    for u, st in dead.items():
        for dev in ("PC", "MO"):
            row = {"url": u, "device": dev, "sitecode": get_site_type(u), "result": "error_page",
                   "final_url": u, "http_status": st, "title": "",
                   "detail": f"HTTP {st} (사전 확인, 브라우저 미사용)", "elapsed": 0.0}
            rows.append(row)
            results.setdefault(u, {})[dev] = "error_page"
            _append_row(row)

    # 일시적 실패(네트워크/타임아웃/세션 끊김)만 재시도. 404·리다이렉트 등 "판정 결과"는 재시도해도 같다.
    RETRYABLE = {"timeout", "error", "mhtml_failed"}

    def _run(task):
        u, dev = task
        _mark_task_start(u, dev)          # 데드라인 감시 대상에 등록
        try:
            info = capture_page(u, dev)
            for attempt in range(1, RETRY_COUNT + 1):
                if info["result"] not in RETRYABLE:
                    break
                print(f"  🔁 재시도 {attempt}/{RETRY_COUNT} ({info['result']}): {dev} {u}")
                drop_thread_driver()      # 깨진 세션을 물고 재시도하지 않도록
                _mark_task_start(u, dev)  # 재시도는 시간을 새로 잰다
                info = capture_page(u, dev)
                info["detail"] = (info["detail"] + f" | retry {attempt}").strip(" |")
        finally:
            _mark_task_end()
        _append_row(info)                 # 끝날 때까지 기다리지 않고 즉시 디스크에 남긴다
        with lock:
            progress["n"] += 1
            # 스레드 로그가 섞이므로 완료 라인만 lock 으로 묶어 깔끔히 출력
            print(f"  ✔ [{progress['n']}/{len(tasks)}] {dev} {u} → {info['result']}")
            # 중간 저장 — 프로세스가 죽어도 여기까지는 리포트에 남는다 (조용히)
            if REPORT_FLUSH_EVERY and progress["n"] % REPORT_FLUSH_EVERY == 0:
                try:
                    write_daily_report(csv_path, ts.split("_")[0], quiet=True)
                except Exception as e:
                    print(f"  ⚠️ 리포트 중간 저장 실패(계속 진행): {e}")
        return info

    # 상한을 넘긴 태스크를 끊어 워커가 영구히 잡히는 것을 막는다
    _stop_watchdog = threading.Event()
    start_deadline_watchdog(_stop_watchdog)

    try:
        with ThreadPoolExecutor(max_workers=max_workers) as ex:
            futures = [ex.submit(_run, t) for t in tasks]
            for fut in as_completed(futures):
                try:
                    info = fut.result()
                except Exception as e:
                    # capture_page 안에서 다 잡히지만, 예기치 못한 경우까지 행으로 남긴다
                    print(f"  ❌ 태스크 에러: {e}")
                    rows.append({"url": "", "device": "", "sitecode": "", "result": "task_error",
                                 "final_url": "", "http_status": "", "title": "",
                                 "detail": f"{type(e).__name__}: {e}"[:200], "elapsed": 0.0})
                    continue
                rows.append(info)
                results.setdefault(info["url"], {})[info["device"]] = info["result"]
    finally:
        _stop_watchdog.set()
        # 재사용하던 Chrome 을 모두 닫는다 (중단·예외로 빠져나가도 좀비가 안 남게)
        quit_all_drivers()

    # 입력 순서대로 skip / error 집계 (PC·MO 중 하나라도 해당되면 1번만 기록)
    skipped_urls = []     # 리다이렉트로 skip된 URL
    error_page_urls = []  # 에러 페이지로 skip된 URL
    unknown_page_urls = []  # unknown(soft-404, og:url 없음) 으로 skip된 URL
    login_page_urls = []  # 로그인/인증 화면으로 skip된 URL
    for u in urls:
        vals = tuple(results.get(u, {}).values())
        if "redirected" in vals:
            skipped_urls.append(u)
        elif "login_page" in vals:
            login_page_urls.append(u)
        elif "error_page" in vals:
            error_page_urls.append(u)
        elif "unknown_page" in vals:
            unknown_page_urls.append(u)

    # ── (url, device) 단위 상세 결과 CSV ──────────────────────
    # txt 3종은 URL 단위라 "PC 는 정상인데 MO 만 리다이렉트" 같은 경우를 구분 못 하고,
    # error/timeout 은 아예 어느 txt 에도 안 남아 재실행 대상 파악이 불가능했다.
    # ⚠ ts / csv_path / _cols 는 캡처 시작 전에 이미 만들어 건별로 append 해 왔다.
    #   여기서는 같은 파일을 "입력 URL 순서"로 정렬해 다시 쓴다(완주했을 때만 정렬본이 남음).
    _order = {u: i for i, u in enumerate(urls)}
    rows.sort(key=lambda r: (_order.get(r.get("url", ""), 10**9), r.get("device", "")))
    with open(csv_path, "w", encoding="utf-8-sig", newline="") as cf:
        w = csv.DictWriter(cf, fieldnames=_cols, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)
    _cnt = Counter(r.get("result", "") for r in rows)
    print(f"📄 결과 CSV {len(rows)}행 → {csv_path}")
    print("   " + " / ".join(f"{k}:{v}" for k, v in sorted(_cnt.items())))

    # ── skip된 URL 요약(콘솔) + 선택적 txt 저장 ───────────────
    # 개수는 항상 콘솔에 찍고, 파일은 WRITE_SKIPPED_TXT 일 때만 남긴다
    # (같은 정보가 _daily_report.xlsx '이슈' 열에 이미 들어간다).
    for label, urls_list, key in (
        ("⏭️  리다이렉트 skip", skipped_urls,      "redirect"),
        ("⛔ 에러 페이지 skip", error_page_urls,   "error_page"),
        ("🔒 로그인 화면 skip", login_page_urls,   "login_page"),
        ("⛔ unknown 페이지 skip", unknown_page_urls, "unknown_page"),
    ):
        if not urls_list:
            continue
        if WRITE_SKIPPED_TXT:
            p = f"{OUTPUT_DIR}/{SKIPPED_TXT_NAMES[key]}"
            with open(p, "w", encoding="utf-8") as f:
                f.write("\n".join(urls_list) + "\n")
            print(f"{label} {len(urls_list)}개 → {p}")
        else:
            print(f"{label} {len(urls_list)}개")

    # ── 일일 리포트 (최신 날짜가 항상 B~D열) ──────────────────
    write_daily_report(csv_path, ts.split("_")[0])

    # ── 완주했으니 결과 CSV 를 치운다(폴더엔 _daily_report.xlsx 만 남긴다) ──
    # 리포트 갱신을 마친 뒤이므로 이력은 리포트가 이미 갖고 있다. 크래시 시엔 여기 못 와서 CSV 가 남는다.
    if not KEEP_RESULT_CSV:
        try:
            os.remove(csv_path)
        except OSError as e:
            print(f"  ⚠️ 결과 CSV 삭제 실패(무시): {e}")

    print("✨ 모든 캡처 완료!")


if __name__ == "__main__":
    capture_urls(URLS)
